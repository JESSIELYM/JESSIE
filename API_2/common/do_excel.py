from openpyxl import load_workbook
from 接口自动化.API_2.common import project_path
from 接口自动化.API_2.common.test_log import MyLog
from 接口自动化.API_2.common.TestConfig import ReadConfig


class DoExcel():
    '''该类完成测试用例的读取及测试结果的返回'''
    def __init__(self,file,sheet):
        '''将Excel的工作簿、表单作为初始化函数的参数'''
        self.file=file
        self.sheet =sheet
    def Read_Excel(self):
        try:
            case_id=ReadConfig(project_path.config_ini_path).get_data('CASE','case_id')
            wb = load_workbook(self.file)
            sheet=wb[self.sheet]
        except Exception as e:
            MyLog().error('加载Excel出错了，错误信息是{}'.format(e))
        else:
            MyLog().info('开始读取测试用例')
            test_data=[]      #将所有用例存储到列表中
            for item in range(2,sheet.max_row+1):    #通过遍历将用例信息以字典形式存储
                row_data={}
                row_data['CaseID']=sheet.cell(item,1).value
                row_data['Module']=sheet.cell(item,2).value
                row_data['Description'] = sheet.cell(item,3).value
                row_data['Url'] = sheet.cell(item,4).value
                row_data['Method'] = sheet.cell(item,5).value
                row_data['Params'] = sheet.cell(item,6).value
                row_data['ExpectedResult'] = sheet.cell(item,7).value
                row_data['ActualResult'] = sheet.cell(item,8).value
                row_data['TestResult'] = sheet.cell(item,9).value
                test_data.append(row_data)
                # MyLog().info('这是读取的第{}条测试用例,用例内容是{}'.format(row_data['CaseID'],test_data[row_data['CaseID']-1]))
            wb.close()
            MyLog().info('读取测试用例结束,')
            finally_data=[]
            if case_id=='all':
                finally_data=test_data
            else:
                for item in test_data:
                    finally_data.append(item)
            return finally_data

    def Write_back(self,row,col,value):
        '''写回测试结果到Excel中'''
        try:
            wb = load_workbook(self.file)
            sheet=wb[self.sheet]
        except Exception as e:
            MyLog().error('加载Excel出错了，错误信息是{}'.format(e))
        try:
            sheet.cell(row,col).value=value
            wb.save(self.file)
            wb.close()
            MyLog().info('向Excel中写入数据，存入的数据信息是{}'.format(value))
        except Exception as e:
            MyLog().error('写入测试结果出错了，错误信息是{}'.format(e))

if __name__=='__main__':
    file=r'D:\PyCharm 2018.1\PycharmProjects\接口自动化\API_1\test_case\接口自动化测试用例.xlsx'
    sheet='API_01_TC'
    res=DoExcel(file,sheet).Read_Excel()
    # print(res)