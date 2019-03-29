from openpyxl import load_workbook
from 接口自动化.接口自动化阶段一作业.HTTPRequests import HttpRequests
import unittest
from ddt import ddt,data,unpack
import logging
class ImportExcel():
    def __init__(self,filename=r'C:\Users\JESSIE\Desktop\接口自动化测试用例.xlsx',sheetname='RegisterTestCase'):
        '''将Excel的工作簿、表单作为初始化函数的参数'''
        self.file=filename
        self.wb = load_workbook(self.file)
        self.sheet = self.wb[sheetname]
    def Read_Method(self,rownum):
        Method=self.sheet.cell(row=rownum,column=4).value
        return Method
    def Read_Params(self,rownum):
        Param=self.sheet.cell(row=rownum,column=5).value
        return eval(Param)   #使用eval函数，将excel中的数据类型转换为原本的格式
    def Read_ExpectedResult(self,rownum):
        EXresult_1=self.sheet.cell(row=rownum,column=6).value
        return eval(EXresult_1)

@ddt #装饰测试类
class RunTestCase(unittest.TestCase):
    def setUp(self):#开始-->在每一条用例执行之前
        print('----开始执行测试了---')#准备工作/准备测试环境
    def tearDown(self):#结束-->在每一条用例执行结束之后
        print('----测试执行完毕了----')#清场工作/清除测试环境
    @data(1,2,3,4,5,6,7,8)
    def test_1(self,rownum):   #正常注册:输入正确的手机号、密码、用户名
        print('我在执行第{}条用例'.format(rownum))
        rownum+=1
        try:
            Method_1=ImportExcel().Read_Method(rownum)
            param_1=ImportExcel().Read_Params(rownum)
            url_1='http://47.107.168.87:8080/futureloan/mvc/api/member/register'
            EXresult_1=ImportExcel().Read_ExpectedResult(rownum)
            ACresult_1=HttpRequests(url_1,Method_1,param_1).requests()
            print(type(ACresult_1))
            print('''
                请求方式为：{}
                输入参数为：{}
                期望结果为：{}
                实际结果为：{}'''.format(Method_1,param_1,EXresult_1,ACresult_1))
        except AttributeError as e:
            print('从excel读取数据出错了，错误为：{}'.format(e))
        try:
            write_excel = ImportExcel()
            write_excel.sheet.cell(row=rownum, column=7, value=str(ACresult_1))
            write_excel.wb.save(ImportExcel().file)  # 保存excel
            write_excel.wb.close()    #关闭
        except Exception as e:
            print('excel写入数据报错:错误信息是{}'.format(e))
        try:
            EXcode=EXresult_1['code']
            ACcode=ACresult_1['code']
            print(EXcode,ACcode)
            self.assertEqual(EXcode,ACcode)  # 设置断言将期望结果和实际结果进行比对
        except AssertionError as e:
            print('用例执行失败，错误是{}'.format(e))
            raise e  # 处理错误之后，需继续抛出异常，才能识别失败的用例
        try:
            write_excel = ImportExcel()
            write_excel.sheet.cell(row=rownum, column=7, value=str(ACresult_1))
            if self.assertEqual(EXcode,ACcode)!='fail':      #如果期望与实际值相等，则写入pass
                write_excel.sheet.cell(row=rownum, column=8, value='Pass')
            else:                                              #期望与实际不等，则写入Failed---------------------------此处不知道为什么无法写入到excel里
                write_excel.sheet.cell(row=rownum, column=8, value='Failed')
            write_excel.wb.save(ImportExcel().file)  # 保存时，要确定excel是关闭状态，不然会报错。
            write_excel.wb.close()
        except Exception as e:
            print('excel写入数据报错:错误信息是{}'.format(e))
    @data(9,10,11,12,13)
    def test_2(self,rownum):   #正常注册:输入正确的手机号、密码、用户名
        print('我在执行第{}条用例'.format(rownum))
        rownum+=1
        try:
            Method_1=ImportExcel().Read_Method(rownum)
            param_1=ImportExcel().Read_Params(rownum)
            url_1='http://47.107.168.87:8080/futureloan/mvc/api/member/login'
            EXresult_1=ImportExcel().Read_ExpectedResult(rownum)
            ACresult_1=HttpRequests(url_1,Method_1,param_1).requests()
            print(type(ACresult_1))
            print('''
                    请求方式为：{}
                    输入参数为：{}
                    期望结果为：{}
                    实际结果为：{}'''.format(Method_1,param_1,EXresult_1,ACresult_1))
        except AttributeError as e:
            print('从excel读取数据出错了，错误为：{}'.format(e))
        try:
            EXcode=EXresult_1['code']
            ACcode=ACresult_1['code']
            print(EXcode,ACcode)
            self.assertEqual(EXcode,ACcode)  # 设置断言将期望结果和实际结果进行比对
        except AssertionError as e:
            print('用例执行失败，错误是{}'.format(e))
            raise e  # 处理错误之后，需继续抛出异常，才能识别失败的用例
        try:
            write_excel=ImportExcel()
            write_excel.sheet.cell(row=rownum,column=7,value=str(ACresult_1))
            if self.assertEqual(EXcode,ACcode)!='fail':
                write_excel.sheet.cell(row=rownum, column=8, value='Pass')
            else:
                write_excel.sheet.cell(row=rownum, column=8, value='Failed')
            write_excel.wb.save(ImportExcel().file)  # 保存时，要确定excel是关闭状态，不然会报错。
            write_excel.wb.close()
        except Exception as e:
            print('excel写入数据报错:错误信息是{}'.format(e))



