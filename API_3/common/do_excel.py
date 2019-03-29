from openpyxl import load_workbook
from 接口自动化.API_3.common import project_path
from 接口自动化.API_3.common.test_log import MyLog
from 接口自动化.API_3.common.TestConfig import ReadConfig

class DoExcel():
    '''该类完成测试用例的读取及测试结果的返回'''
    def __init__(self,file,sheet):
        '''将Excel的工作簿、表单作为初始化函数的参数'''
        self.file=file
        self.sheet =sheet
        self.Telsheet='Tel'
    def Read_Excel(self,section):    #section配置文件里的片段名，根据section决定不同表单所执行的具体用例
        try:
            case_id=ReadConfig(project_path.config_ini_path).get_data(section,'case_id')
            wb = load_workbook(self.file)
            sheet=wb[self.sheet]
            Telsheet=wb[self.Telsheet]
        except Exception as e:
            MyLog().error('加载Excel出错了，错误信息是{}'.format(e))
        else:
            MyLog().info('开始读取测试用例')
            test_data=[]      #将所有用例存储到列表中
            for item in range(2,sheet.max_row+1):    #通过遍历将用例信息以字典形式存储
                row_data={}
                row_data['CaseID']=sheet.cell(item,1).value
                row_data['Module']=sheet.cell(item,2).value
                row_data['Description'] = sheet.cell(item,3).value
                row_data['Url'] = sheet.cell(item,4).value
                row_data['Method'] = sheet.cell(item,5).value
                # if 'tel' in sheet.cell(item, 6).value:    #也可使用成员运算符判断参数中是否包含有tel这个字符串
                try:
                    if sheet.cell(item, 6).value.find('tel')!=-1:    #使用find查找参数中是否有tel这个字符串，有的话就替换成正确的手机号码，对手机号码做参数化
                        row_data['Params'] = sheet.cell(item, 6).value.replace('tel',str(Telsheet.cell(2,1).value))
                        Telsheet.cell(2, 1).value += 1   #更新tel的初始值
                        wb.save(self.file)
                    else:
                        row_data['Params'] = sheet.cell(item, 6).value
                except Exception as e:
                    MyLog().error('手机号替换不成功，失败原因是{}'.format(e))
                finally:
                    # row_data['Params'] = sheet.cell(item, 6).value
                    row_data['ExpectedResult'] = sheet.cell(item,7).value
                    row_data['ActualResult'] = sheet.cell(item,8).value
                    row_data['TestResult'] = sheet.cell(item,9).value
                    test_data.append(row_data)
                # MyLog().info('这是读取的第{}条测试用例,用例内容是{}'.format(row_data['CaseID'],test_data[row_data['CaseID']-1]))
            wb.close()
            MyLog().info('读取测试用例结束,')
            finally_data=[]
            if case_id=='all':
                finally_data=test_data
            else:
                for item in test_data:
                    finally_data.append(item)
            return finally_data

    def Write_back(self,row,col,value):
        '''写回测试结果到Excel中'''
        try:
            wb = load_workbook(self.file)
            sheet=wb[self.sheet]
            Telsheet=wb[self.Telsheet]
        except Exception as e:
            MyLog().error('加载Excel出错了，错误信息是{}'.format(e))
        try:
            sheet.cell(row,col).value=value
            Telsheet.cell(2,1).value+=1
            wb.save(self.file)
            wb.close()
            MyLog().info('向Excel中写入数据，存入的数据信息是{}'.format(value))
        except Exception as e:
            MyLog().error('写入测试结果出错了，错误信息是{}'.format(e))

if __name__=='__main__':
    file=r'D:\PyCharm 2018.1\PycharmProjects\接口自动化\API_3\test_case\接口自动化测试用例.xlsx'
    sheet='register'
    res=DoExcel(file,sheet).Read_Excel('RegisterCASE')
    print(res)